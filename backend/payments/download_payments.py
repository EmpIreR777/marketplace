from datetime import datetime
import re
from django.utils.timezone import localtime
from django.db.models import QuerySet
from django.http import HttpResponse

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from courses.models import Course


def clean_string(value):
    if isinstance(value, str):
        value = re.sub(r"[^\w\s@.+-]", "", value)
        value = value.strip()
    return value


def export_payments_excel(queryset: QuerySet):
    date_now = datetime.now()
    date = date_now.strftime("%d.%m.%Y %H:%M:%S")
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="payments_report_{date}.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments Report"

    headers = ["ID платежа", "Курс", "Автор", "Email автора", "Телефон автора", 
               "Покупатель", "Email покупателя", "Телефон покупателя",
               "Сумма", "Тип оплаты", "Статус", "Дата"]
    ws.append(headers)

    for payment in queryset:
        course = Course.objects.get(id=payment.item_id)
        author = course.author.title
        phone_author = clean_string(course.author.phone_number) if course.author.phone_number else ""
        email_author = clean_string(course.author.email) if course.author.email else ""
        student_name = f"{payment.user.last_name} {payment.user.first_name} {payment.user.middle_name}"
        email_student = clean_string(payment.user.email) if payment.user.email else ""
        phone_student = clean_string(payment.user.phone_number) if payment.user.phone_number else ""
        amount = f"{payment.amount} руб"
        created_at = localtime(payment.created_at).strftime("%d.%m.%Y %H:%M")

        row = [
            payment.id,
            course.name,
            author,
            email_author,
            phone_author,
            student_name,
            email_student,
            phone_student,
            amount,
            payment.get_payment_type_display(),
            payment.get_status_display(),
            created_at
        ]
        ws.append(row)


    for col in range(1, len(headers) + 1):
        column = get_column_letter(col)
        max_length = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response
